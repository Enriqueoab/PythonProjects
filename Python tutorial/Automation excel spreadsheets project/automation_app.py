import openpyxl as xl
from openpyxl.chart import BarChart, Reference

# Function to reduce the value of a column in a sheet and save it in another column


def reduce_price(file_url):
    # Function for loading a xl
    wb = xl.load_workbook(file_url)
    # Access to the sheet of our document
    sheet = wb['Sheet1']
    # Ways to access to a cell
    cell = sheet['a1']
    cell = sheet.cell(1, 1)

    # print(cell.value) #Value of a cell
    # print(sheet.max_row)
    sheet.cell(1, sheet.max_column).value = 'Corrected price'  # Set a title for the last column

    # Loop to get the data from the columns
    for row in range(2, sheet.max_row + 1):
        cell=sheet.cell(row, 3)  # Get the cell
        corrected_price = cell.value * 0.9  # Get the value and modify
        corrected_price_cell = sheet.cell(row, 4)  # Get the cell where the new value it is going to be
        corrected_price_cell.value = corrected_price  # Set the new value in the cell
    wb.save(file_url)  # Save the document in a new one

# Function to add a chart to our sheet


def set_a_chart(file_url):
    wb = xl.load_workbook(file_url)
    sheet = wb['Sheet1']

    fourth_col_values = Reference(sheet,
              min_row = 2,  # To avoid take the 1 row which are the titles
              max_row = sheet.max_row,  # Take all tha values
              min_col = 4,    # Limited to the data from the fourth column only
              max_col = 4)    # Limited to the data from the fourth column only

    chart = BarChart()  # Intance of the object BarChar
    chart.add_data(fourth_col_values)  # Set the data that the chart is going to show
    sheet.add_chart(chart, 'a8')    # Set the chart to the sheet and its position
    wb.save(file_url)  # Save the document in a new one


reduce_price('transactions.xlsx')
set_a_chart('transactions.xlsx')
